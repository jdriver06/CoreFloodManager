#!/usr/bin/env python

__author__ = 'Jonathan W. Driver, PhD'
__version__ = '1.3.0'
# Updates from 1.0.0: (1) sheet selector, (2) import of 2nd format, (3) Li+ and F-, (4) multi-source, (5) 1x, 2x, etc.,
# (6) selector clearing, (7) alphabetical ordering, (8) fix error if no brine imported, (9) automatic commit when
# sending to database
# Updates from 1.1.0: (1) additive tool for additive salts, e.g. Na2CO3, (2) PrintView to see salt composition
# Updates from 1.2.0: (1) NaCl a new additive salt, (2) UI update triggered when concentration or total mass adjusted

import sqlite3 as sql
from math import isnan, exp
import tkinter as tk
from tkinter.filedialog import askopenfile
from xlrd import open_workbook
from openpyxl import load_workbook
from os.path import splitext, isdir
from dymo import print_to_dymo
from datetime import datetime
from functools import partial
import sys

db_default_path = 'U:\\UEORLAB1\\Python Files\\'
# db_default_path = 'C:\\Users\\jdriver\\Dropbox\\Public\\PycharmProjects\\BrineDatabase\\'
# db_default_path = 'C:\\Users\\Jonathan\\Dropbox\\Public\\PycharmProjects\\BrineDatabase\\'
db_default_name = 'brine_database v1-3-0.db'

gui_width = 662 + 170
gui_height = 19 * 25 - 12


class OptionDialog(tk.Toplevel):

    def __init__(self, parent, title: str, question: str, options: list):
        tk.Toplevel.__init__(self, parent)
        self.title(title)
        self.question = question
        self.transient(parent)
        self.protocol('WM_DELETE_WINDOW', self.cancel)
        self.result = '_'

        self.listbox = tk.Listbox(self)
        self.listbox.bind('<Double-Button-1>', self.set_option)
        for option in options:
            self.listbox.insert(tk.END, option)
        self.listbox.pack(expand=1, fill=tk.BOTH)

    def set_option(self, _):
        self.result = self.listbox.curselection()[0]
        self.destroy()

    def cancel(self):
        self.result = None
        self.destroy()


class Brine:

    ion_names = {"Li+": 'lithium', "Na+": 'sodium', "K+": 'potassium', "Mg++": 'magnesium', "Ca++": 'calcium',
                 "Ba++": 'barium', "Sr++": 'strontium', "Fe++": 'ironII', "F-": 'fluoride', "Cl-": 'chloride',
                 "Br-": 'bromide', "HCO3-": 'bicarbonate', "CO3--": 'carbonate', "SO4--": 'sulfate'}

    ion_weights = {"Li+": 6.94, "Na+": 22.99, "K+": 39.1, "Mg++": 24.305, "Ca++": 40.078, "Ba++": 137.327,
                   "Sr++": 87.62, "Fe++": 55.84, "F-": 19.0, "Cl-": 35.453, "Br-": 79.904, "HCO3-": 61.01,
                   "CO3--": 60.008, "SO4--": 96.06}

    def __init__(self, lithium=0., sodium=0., potassium=0., magnesium=0., calcium=0., barium=0.,
                 strontium=0., iron2=0., fluoride=0., chloride=0., bromide=0., bicarbonate=0., carbonate=0.,
                 sulfate=0.):

        self.composition = {"Li+": lithium, "Na+": sodium, "K+": potassium, "Mg++": magnesium, "Ca++": calcium,
                            "Ba++": barium, "Sr++": strontium, "Fe++": iron2, "F-": fluoride, "Cl-": chloride,
                            "Br-": bromide, "HCO3-": bicarbonate, "CO3--": carbonate, "SO4--": sulfate}
        self.name = ''

    def set_ion(self, ion, value):
        if ion in self.composition.keys():
            self.composition[ion] = value

    def add_ion(self, ion, value):
        if ion in self.composition.keys():
            self.composition[ion] += value

    def get_tds(self):
        tds = 0.
        for value in self.composition.values():
            tds += value
        return tds

    def get_viscosity(self, temp):

        # convert temp to Celsius, calculate TDS in ppk
        # temp = 5. * (temp - 32.) / 9.
        sal = 1.0e-3 * self.get_tds()

        # second-order model parameters for relative viscosity vs. temperature
        a = 1.474e-3 + 1.5e-5 * temp + 3.927e-8 * temp ** 2
        b = 1.0734e-5 - 8.5e-8 * temp + 2.23e-10 * temp ** 2
        mur = 1 + a * sal + b * sal ** 2
        # model of DI viscosity vs. temperature
        muw = exp(-3.79418 + 604.129 / (139.18 + temp))

        return muw * mur


class NominalBrine(Brine):

    source_salts = {"LiCl": 42.39, "NaCl": 58.443, "KCl": 74.553, "CaCl2_2H2O": 147.01, "MgCl2_6H2O": 203.31,
                    "BaCl2": 208.23, "SrCl2": 158.526, "FeCl2": 126.751, "NaF": 41.99, "NaHCO3": 84.007,
                    "Na2SO4": 142.04, "NaBr": 102.894, "SrCl2_6H2O": 266.64, "BaCl2_2H2O": 244.26}
    unique_sources = {"Li+": "LiCl", "K+": "KCl", "Ca++": "CaCl2_2H2O", "Mg++": "MgCl2_6H2O",
                      "Ba++": ["BaCl2", "BaCl2_2H2O"], "Fe++": "FeCl2", "Sr++": ["SrCl2", "SrCl2_6H2O"], "F-": "NaF",
                      "HCO3-": "NaHCO3", "SO4--": "Na2SO4", "Br-": "NaBr"}
    additive_salts = {"Na2CO3": [[2, 1], ["Na+", "CO3--"], 105.989], "NaCl": [[1, 1], ["Na+", "Cl-"], 58.443]}

    def __init__(self, lithium=0., sodium=0., potassium=0., magnesium=0., calcium=0., barium=0.,
                 strontium=0., iron2=0., fluoride=0., chloride=0., bromide=0., bicarbonate=0., carbonate=0.,
                 sulfate=0., sodium_bicarbonate=0., sodium_chloride=0.):

        Brine.__init__(self)
        self.nominal_composition = {"Li+": lithium, "Na+": sodium, "K+": potassium, "Mg++": magnesium, "Ca++": calcium,
                                    "Ba++": barium, "Sr++": strontium, "Fe++": iron2, "F-": fluoride, "Cl-": chloride,
                                    "Br-": bromide, "HCO3-": bicarbonate, "CO3--": carbonate, "SO4--": sulfate}
        self.add_salt_composition = {"Na2CO3": sodium_bicarbonate, "NaCl": sodium_chloride}
        self.salt_composition = {}
        self.source_options = {"Ba++": 0, "Sr++": 0}
        self.update_composition()

    def update_composition(self):

        self.salt_composition = {}
        ss = NominalBrine.source_salts
        for key, value in NominalBrine.unique_sources.items():
            if type(value) == list:
                value = value[self.source_options[key]]
            ion_ppm = self.nominal_composition[key]
            self.salt_composition[value] = ion_ppm * ss[value] / Brine.ion_weights[key]
        na_ppm = self.nominal_composition["Na+"]
        na_w = Brine.ion_weights["Na+"]

        self.salt_composition["NaCl"] = (na_ppm / na_w) * ss["NaCl"] - \
                                        (ss["NaCl"] / ss["NaHCO3"]) * self.salt_composition["NaHCO3"] - \
                                        (2. * ss["NaCl"] / ss["Na2SO4"]) * self.salt_composition["Na2SO4"] - \
                                        (ss["NaCl"] / ss["NaBr"]) * self.salt_composition["NaBr"] - \
                                        (ss["NaCl"] / ss["NaF"]) * self.salt_composition["NaF"]
        cl_ppm = 0.
        for key, value in self.salt_composition.items():
            if key.find("Cl") != -1:
                d = 1.
                if key.find("Cl2") != -1:
                    d = 2.
                cl_ppm += d * value * Brine.ion_weights["Cl-"] / NominalBrine.source_salts[key]
        self.set_ion("Cl-", cl_ppm)
        self.set_ion("Na+", na_ppm)
        self.set_ion("CO3--", 0)
        for key in self.unique_sources.keys():
            self.set_ion(key, self.nominal_composition[key])
        for key, value in self.add_salt_composition.items():
            info = NominalBrine.additive_salts[key]
            mm = info[2]
            for valency, ion in zip(info[0], info[1]):
                ion_mm = Brine.ion_weights[ion]
                ion_add_ppm = int((valency * ion_mm / mm) * value * 10000)
                self.add_ion(ion, ion_add_ppm)

    def set_add_salt_composition(self, add_salt: str, value: int):
        self.add_salt_composition[add_salt] = value

    def get_n_tds(self):
        n_tds = 0.
        for value in self.nominal_composition.values():
            n_tds += value
        return n_tds


class TableEntry(tk.Entry):

    def __init__(self, parent, row: int, col: int, entry_width: int):
        tk.Entry.__init__(self, parent)
        self.row = row
        self.col = col
        self.configure(justify=tk.CENTER, width=entry_width)
        self.bind('<Key>', parent.key_press)
        self.bind('<Button-1>', parent.mouse_press)
        self.bind('<Triple-Button-1>', parent.fill_down)
        self.bind('<FocusOut>', parent.focus_out)
        self.bind('<FocusIn>', parent.check_focus)
        self.var = tk.StringVar(self)
        self.var.set('')
        self.config(textvariable=self.var)


class AskSheetOption(OptionDialog):

    def __init__(self, parent, sheet_names):
        OptionDialog.__init__(self, parent, 'Select Sheet', '', sheet_names)
        self.parent = parent
        self.sheet_names = sheet_names

    def destroy(self):
        self.parent.geometry(str(gui_width) + 'x' + str(19*25))
        if self.result is not None:
            if self.parent.brine_view.import_type == '.xls':
                ws = self.parent.brine_view.wb.sheet_by_name(self.sheet_names[self.result])
            else:
                ws = self.parent.brine_view.wb.get_sheet_by_name(self.sheet_names[self.result])
            self.parent.brine_view.read_and_load_data(ws)
        OptionDialog.destroy(self)


class AskSaltSourceOption(OptionDialog):

    def __init__(self, parent, ion, salt_names):
        OptionDialog.__init__(self, parent, 'Select Salt', '', salt_names)
        self.parent = parent
        self.salt_names = salt_names
        self.ion = ion

    def destroy(self):
        self.parent.geometry(str(gui_width) + 'x' + str(21*25))
        if self.result is not None:
            self.parent.brine_view.n_brine.source_options[self.ion] = self.result
            self.parent.brine_view.n_brine.update_composition()
            sorted_print_list = self.parent.brine_view.create_print_list
            self.parent.print_view.display_print_list(sorted_print_list)
        OptionDialog.destroy(self)


class BrineView(tk.Frame):

    def __init__(self, parent, nb: NominalBrine):
        tk.Frame.__init__(self, parent)
        self.parent = parent
        self.n_brine = nb
        self.m_or_k = 'k'
        self.import_type = '_'
        self.wb = None

        name_label = tk.Label(self, text='Brine Name:')
        self.name_entry = tk.Entry(self)
        self.name_entry_var = tk.StringVar(self)
        self.name_entry_var.set('')
        self.name_entry.config(textvariable=self.name_entry_var, width=47)
        name_label.grid(row=0, column=0, pady=(5, 0))
        self.name_entry.grid(row=0, column=1, columnspan=4, pady=(5, 0))

        abbrev_label = tk.Label(self, text='Abbreviation:')
        self.abbrev_entry = tk.Entry(self)
        self.abbrev_entry_var = tk.StringVar(self)
        self.abbrev_entry_var.set('')
        self.abbrev_entry.config(textvariable=self.abbrev_entry_var, width=47)
        abbrev_label.grid(row=1, column=0)
        self.abbrev_entry.grid(row=1, column=1, columnspan=4)

        self.headers = [tk.Label(self, text='Ion'), tk.Label(self, text='Target [ppm]'),
                        tk.Label(self, text='Actual [ppm]'), tk.Label(self, text='  Additive  '),
                        tk.Label(self, text='%')]
        j = 0
        for header in self.headers:
            header.grid(row=2, column=j, pady=(4, 0))
            j += 1

        i = 3
        self.labels = []
        self.entries = []
        self.actual_entries = []
        for key, value in nb.nominal_composition.items():
            label = tk.Label(self, text=key)
            if key in NominalBrine.unique_sources.keys():
                if type(NominalBrine.unique_sources[key]) == list:
                    label.bind('<Enter>', self.ion_option_highlight)
                    label.bind('<Leave>', self.ion_option_unhighlight)
                    label.bind('<Button-1>', self.ask_source_option)
            label.grid(row=i, column=0)
            self.labels.append(label)
            entry = TableEntry(self, i, 0, 10)
            entry.var.set(str(int(value)))
            self.entries.append(entry)
            entry.grid(row=i, column=1)
            if key == 'CO3--':
                entry.var.set('--')
                entry.config(state='disabled')
            a_entry = tk.Entry(self)
            a_entry.grid(row=i, column=2)
            a_entry.config(width=10, justify=tk.CENTER, state='disabled')
            self.actual_entries.append(a_entry)
            i += 1

        j = 3
        self.add_labels = []
        self.add_entries = []
        for key, value in nb.add_salt_composition.items():
            label = tk.Label(self, text=key)
            label.grid(row=j, column=3)
            self.add_labels.append(label)
            entry = TableEntry(self, j, 1, 10)
            entry.var.set(str(float(value)))
            self.add_entries.append(entry)
            entry.grid(row=j, column=4)
            j += 1

        self.tds_label = tk.Label(self, text='TDS')
        self.tds_label.grid(row=i, column=0)
        self.tds_entry = tk.Entry(self)
        self.tds_entry.grid(row=i, column=1)
        self.tds_entry.insert(0, str(int(nb.get_n_tds())))
        self.tds_entry.config(width=10, justify=tk.CENTER, state='disabled')
        self.actual_tds_entry = tk.Entry(self)
        self.actual_tds_entry.grid(row=i, column=2)
        self.actual_tds_entry.config(width=10, justify=tk.CENTER, state='disabled')

        self.button_frame = tk.Frame(self)
        self.import_button = tk.Button(self.button_frame, text='Import', command=self.import_command)
        self.import_button.config(width=28, state='disabled')
        self.import_button.pack(side=tk.TOP, padx=0)
        self.send_button = tk.Button(self.button_frame, text='Send', command=self.send_command)
        self.send_button.config(width=28, state='disabled')
        self.send_button.pack(side=tk.TOP, padx=0)
        self.button_frame.grid(row=i+2, column=0, rowspan=2, columnspan=3, pady=(7, 5))

        self.total_mass_label = tk.Label(self, text='Total Mass [g]:')
        self.total_mass_entry = tk.Entry(self)
        self.total_mass_entry_var = tk.StringVar(self)
        self.total_mass_entry_var.set('4000')
        self.total_mass_entry.config(textvariable=self.total_mass_entry_var, width=10)
        self.total_mass_label.grid(row=i+2, column=3, pady=(10, 0))
        self.total_mass_entry.grid(row=i+2, column=4, pady=(10, 0))
        self.total_mass_entry.bind('<FocusOut>', self.create_and_display_print_list)
        self.total_mass_entry.bind('<Return>', self.create_and_display_print_list)

        self.conc_factor_label = tk.Label(self, text='Conc. Factor:')
        self.conc_factor_entry = tk.Entry(self)
        self.conc_factor_entry_var = tk.StringVar(self)
        self.conc_factor_entry_var.set('1')
        self.conc_factor_entry.config(textvariable=self.conc_factor_entry_var, width=10)
        self.conc_factor_label.grid(row=i+3, column=3, pady=(0, 10))
        self.conc_factor_entry.grid(row=i+3, column=4, pady=(0, 10))
        self.conc_factor_entry.bind('<FocusOut>', self.create_and_display_print_list)
        self.conc_factor_entry.bind('<Return>', self.create_and_display_print_list)

        # self.print_button = tk.Button(self, text='Print', command=self.print_command)
        # self.print_button.config(width=8, height=2)
        # self.print_button.grid(row=i+2, column=3, rowspan=2, columnspan=2, pady=(15, 10))

        self.config(highlightbackground='black', highlightthickness=1)
        self.grid(row=0, column=1, rowspan=1, columnspan=2, padx=(5, 10), ipadx=5, pady=(5, 5))

    def key_press(self, event: tk.EventType.KeyPress):

        self.m_or_k = 'k'
        entry = event.widget
        entry.config(background='Yellow')
        if event.keysym == 'Return' or event.keysym == 'Tab':
            entry.config(background='White')
            if entry in self.entries:
                ion = self.labels[entry.row - 3].cget('text')
                if ion != 'CO3--':
                    try:
                        value = int(entry.var.get())
                        if value < 0:
                            raise ValueError
                        self.n_brine.nominal_composition[ion] = value
                        self.n_brine.update_composition()

                    except ValueError as e:
                        print(e)
                        entry.var.set(str(int(self.n_brine.nominal_composition[ion])))
            elif entry in self.add_entries:
                add_salt = self.add_labels[entry.row - 3].cget('text')
                try:
                    value = float(entry.var.get())
                    if value < 0:
                        raise ValueError
                    self.n_brine.add_salt_composition[add_salt] = value
                    self.n_brine.update_composition()
                except ValueError as e:
                    print(e)
                    entry.var.set(str(int(self.n_brine.add_salt_composition[add_salt])))

            self.update_actual_entries()
            entry.config(background='White')
            self.create_and_display_print_list()

    def update_actual_entries(self):

        j = 0
        for ae in self.actual_entries:
            ae_ion = self.labels[j].cget('text')
            ae.config(state='normal')
            ae.delete(0, tk.END)
            ae.insert(0, str(int(self.n_brine.composition[ae_ion])))
            ae.config(state='disabled')
            j += 1
        self.tds_entry.config(state='normal')
        self.tds_entry.delete(0, tk.END)
        self.tds_entry.insert(0, str(int(self.n_brine.get_n_tds())))
        self.tds_entry.config(state='disabled')
        self.actual_tds_entry.config(state='normal')
        self.actual_tds_entry.delete(0, tk.END)
        self.actual_tds_entry.insert(0, str(int(self.n_brine.get_tds())))
        self.actual_tds_entry.config(state='disabled')

    @staticmethod
    def ion_option_highlight(event: tk.EventType.Enter):
        event.widget.config(fg='blue')

    @staticmethod
    def ion_option_unhighlight(event: tk.EventType.Leave):
        event.widget.config(fg='black')

    def ask_source_option(self, event: tk.EventType.ButtonPress):
        event.widget.config(fg='black')
        ion = event.widget.cget('text')
        salt_names = NominalBrine.unique_sources[ion]
        aso = AskSaltSourceOption(self.parent, ion, salt_names)
        aso.geometry(f'+{self.parent.winfo_x()}+{self.parent.winfo_y()}')
        aso.geometry('150x150')
        self.parent.geometry(str(gui_width) + 'x0')

    def mouse_press(self, event: tk.EventType.ButtonPress):
        self.m_or_k = 'm'
        if event:
            pass

    def fill_down(self, event: tk.EventType.ButtonPress):
        if event or self:
            pass

    def focus_out(self, event: tk.EventType.FocusOut):
        if event or self:
            pass

    def check_focus(self, event: tk.EventType.FocusIn):
        if event or self:
            pass

    def import_command(self):
        dlg = askopenfile(mode='r')
        if dlg is None:
            return
        _, ext = splitext(dlg.name)

        def ask_which_sheet(parent: tk.Tk, sheet_names: list, get_sheet_func):
            if len(sheet_names) == 1:
                self.read_and_load_data(get_sheet_func(sheet_names[0]))
            else:
                option_dialog = AskSheetOption(parent, sheet_names)
                option_dialog.geometry(f'+{parent.winfo_x()}+{parent.winfo_y()}')
                option_dialog.geometry('150x150')
                parent.geometry(str(gui_width) + 'x0')

        if ext == '.xls':
            self.import_type = ext
            wb = open_workbook(dlg.name)
            self.wb = wb
            ask_which_sheet(self.parent, wb.sheet_names(), wb.sheet_by_name)
        elif ext == '.xlsx':
            self.import_type = ext
            wb = load_workbook(dlg.name)
            self.wb = wb
            ask_which_sheet(self.parent, wb.get_sheet_names(), wb.get_sheet_by_name)
        else:
            return

    def read_and_load_data(self, ws):
        data_values = [[], [], [], [], [], []]
        if self.import_type == '_':
            return
        elif self.import_type == '.xls':
            for i in range(20):
                for ii in range(6):
                    data_values[ii].append(ws.cell_value(i, ii))
        else:
            for i in range(20):
                for ii in range(6):
                    data_values[ii].append(ws.cell(row=i+1, column=ii+1).value)
        j = -1
        jj = -1
        for k in range(10):
            if data_values[0][k] == 'Ion':
                j = k
            elif data_values[0][k] == 'Cation':
                jj = k

        if j != -1:
            i = j + 1
            while isinstance(data_values[0][i], str) and data_values[0][i]:
                ion = data_values[0][i]
                for j, label in enumerate(self.labels):
                    if label.cget('text') == ion or (label.cget('text') == 'SO4--' and ion == 'SO4='):
                        entry = self.entries[j]
                        entry.delete(0, tk.END)
                        if data_values[1][i] is not None and type(data_values[1][i]) is not str \
                                and not isnan(data_values[1][i]):
                            entry.insert(0, str(int(data_values[1][i])))
                        else:
                            entry.insert(0, str(0))
                        entry.focus_set()
                        entry.event_generate('<Return>')
                i += 1
        elif jj != -1:
            ion_name_list = list(Brine.ion_names.values())
            ion_list = list(Brine.ion_names.keys())
            cols = [0, 3]
            for col in cols:
                i = jj + 1
                while isinstance(data_values[col][i], str) and data_values[col][i]:
                    ion_name = data_values[col][i].lower()
                    ion_name = ion_name.strip()
                    if ion_name == 'flouride':
                        ion_name = 'fluoride'
                    ind = ion_name_list.index(ion_name)
                    ion = ion_list[ind]
                    for j, label in enumerate(self.labels):
                        if label.cget('text') == ion or (label.cget('text') == 'SO4--' and ion == 'SO4='):
                            entry = self.entries[j]
                            entry.delete(0, tk.END)
                            if data_values[col+2][i] is not None and type(data_values[col+2][i]) is not str \
                                    and not isnan(data_values[col+2][i]):
                                entry.insert(0, str(int(data_values[col+2][i])))
                            else:
                                entry.insert(0, str(0))
                            entry.focus_set()
                            entry.event_generate('<Return>')
                    i += 1

    def send_command(self):
        name = self.name_entry_var.get()
        if not name.strip():
            print('Name is empty.')
            return
        abbrev = self.abbrev_entry_var.get()
        if not abbrev.strip():
            print('Abbreviation is empty.')
            return
        p_id = self.parent.brine_selector.c.execute('SELECT project_id from projects WHERE project_name=?',
                                                    (self.parent.brine_selector.project_tk_var.get(),)).fetchall()[0][0]
        ion_names_list = ['lithium', 'sodium', 'potassium', 'magnesium', 'calcium', 'barium', 'strontium', 'ironII',
                          'fluoride', 'chloride', 'bromide', 'bicarbonate', 'sulfate']
        values = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
        for key, value in self.n_brine.nominal_composition.items():
            if key != 'CO3--':
                ion = Brine.ion_names[key]
                ind = ion_names_list.index(ion)
                values[ind] = value

        add_names_list = ['Na2CO3']
        add_values = [0.]
        for key, value in self.n_brine.add_salt_composition.items():
            ind = add_names_list.index(key)
            add_values[ind] = float(value)

        self.parent.brine_selector.c.execute('''INSERT INTO brines(project_id, brine_name, brine_abbrev, lithium, 
        sodium, potassium, magnesium, calcium, barium, strontium, ironII, fluoride, chloride, bromide, bicarbonate, 
        sulfate, sodium_carbonate) VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                                             (p_id, name, abbrev, *values, *add_values))

        self.import_button.config(state='disabled')
        self.send_button.config(state='disabled')
        self.parent.brine_selector.conn.commit()
        self.parent.brine_selector.brine_listbox_refresh()

    def create_and_display_print_list(self, *args):
        print(args)

        sorted_print_list = self.create_print_list()
        if sorted_print_list:
            self.parent.print_view.display_print_list(sorted_print_list)

    def create_print_list(self):

        name = self.abbrev_entry_var.get()

        try:
            total_mass = float(self.total_mass_entry_var.get())
            water_mass = total_mass
        except ValueError as e:
            print(self.total_mass_entry_var.get())
            print(e)
            return []
        try:
            conc_factor = float(self.conc_factor_entry_var.get())
        except ValueError as e:
            print(self.conc_factor_entry_var.get())
            print(e)
            return []

        print_list = [name]

        masses = []
        for key, value in self.n_brine.salt_composition.items():
            if key in self.n_brine.add_salt_composition.keys():
                value += 10000. * self.n_brine.add_salt_composition[key]
            grams = 1e-6 * total_mass * value * conc_factor
            water_mass -= grams
            if grams > 0:
                masses.append(grams)
                print_key = key
                if print_key.find('H2O') != -1:
                    print_key = print_key[:-3] + 'W'
                print_list.append('{:>7.3f}'.format(grams) + ' g ' + print_key)

        for key, value in self.n_brine.add_salt_composition.items():
            if key in self.n_brine.salt_composition.keys():
                continue
            grams = 1e-2 * total_mass * value * conc_factor
            water_mass -= grams
            if grams > 0:
                masses.append(grams)
                print_key = key
                if print_key.find('H2O') != -1:
                    print_key = print_key[:-3] + 'W'
                print_list.append('{:>7.3f}'.format(grams) + ' g ' + print_key)

        s_masses = sorted(masses, reverse=True)
        sorted_print_list = [print_list[0]]
        for mass in s_masses:
            ind = masses.index(mass)
            sorted_print_list.append(print_list[ind + 1])
        sorted_print_list.append('{:>7.2f}'.format(water_mass) + ' g ' + 'dH2O')

        return sorted_print_list

    def print_command(self):

        try:
            conc_factor = float(self.conc_factor_entry_var.get())
        except ValueError as e:
            print(self.conc_factor_entry_var.get())
            print(e)
            return

        sorted_print_list = self.create_print_list()

        if not sorted_print_list:
            return

        if len(sorted_print_list) < 14:
            for i in range(14 - len(sorted_print_list)):
                sorted_print_list.append('')
            label_file = 'my.label'
        elif len(sorted_print_list) == 14:
            label_file = 'my.label'
        elif 14 < len(sorted_print_list) < 17:
            for i in range(17 - len(sorted_print_list)):
                sorted_print_list.append('')
            label_file = 'brine_with_additives.label'
        elif len(sorted_print_list) > 17:
            print('print list is too long: ', len(sorted_print_list))
            label_file = ''
        else:
            label_file = 'brine_with_additives.label'

        sorted_print_list.append('DATE: ' + datetime.today().strftime('%Y-%m-%d'))
        if conc_factor != 1.:
            sorted_print_list[0] = str(conc_factor) + 'x' + sorted_print_list[0]
        # print(len(sorted_print_list), label_file, sorted_print_list)
        print_to_dymo(sorted_print_list, label_file)


class BrineSelector(tk.Frame):

    def __init__(self, parent):
        tk.Frame.__init__(self, parent)
        self.parent = parent
        self.select_db_button = tk.Button(self, text='Select DB:', command=self.select_db_command)
        self.company_choices = [' ']
        try:
            path = db_default_path
            if not isdir('U:\\'):
                path = 'Q' + path[1:]
            self.conn = sql.connect(path + db_default_name)
            db_name = db_default_name
            self.c = self.conn.cursor()
            result = self.c.execute('''SELECT company_name FROM companies''').fetchall()
            for company in result:
                self.company_choices.append(company[0])
            self.company_choices.sort()
        except Exception as e:
            db_name = ''
            print(e)
        self.db_name_label = tk.Label(self, text=db_name)
        self.db_name_label.config(width=25)
        self.company_label = tk.Label(self, text='Company:')
        self.company_tk_var = tk.StringVar(self)
        self.company_tk_var.set(' ')
        self.company_option = tk.OptionMenu(self, self.company_tk_var, *self.company_choices,
                                            command=self.company_option_command)
        self.company_option.config(width=25)
        self.project_label = tk.Label(self, text='Project:')
        self.project_tk_var = tk.StringVar(self)
        self.project_tk_var.set(' ')
        self.project_choices = [' ']
        self.project_option = tk.OptionMenu(self, self.project_tk_var, *self.project_choices,
                                            command=self.project_option_command)
        self.project_option.config(width=25, state='disabled')
        self.brine_label = tk.Label(self, text='Brine List:')
        self.brine_listbox = tk.Listbox(self, height=19, width=40, state='disabled')
        self.brine_listbox.bind('<Double-Button-1>', self.browse_brine)

        self.select_db_button.grid(row=0, column=0, pady=(5, 0))
        self.db_name_label.grid(row=0, column=1, pady=(5, 0))
        self.company_label.grid(row=1, column=0)
        self.company_option.grid(row=1, column=1)
        self.project_label.grid(row=2, column=0)
        self.project_option.grid(row=2, column=1)
        self.brine_label.grid(row=3, column=0, pady=(15, 0))
        self.brine_listbox.grid(row=4, column=0, columnspan=2, pady=(2, 12))
        self.brine_list = []

        self.config(highlightbackground='black', highlightthickness=1)
        self.grid(row=0, column=0, rowspan=1, columnspan=1, padx=(10, 5), pady=(5, 5))
        self.select_db_button.config(state='disabled')

    def select_db_command(self):
        print(self)

    def company_option_command(self, *args):
        company = args[0]
        self.project_choices = [' ']
        self.project_option.config(state='normal')
        if company == ' ':
            self.project_choices = [' ']
            self.refresh_project_option()
            self.project_tk_var.set(' ')
            self.project_option.config(state='disabled')
            self.brine_listbox.delete(0, tk.END)
            self.brine_listbox.config(state='disabled')
        else:
            result = self.c.execute('SELECT company_id FROM companies WHERE company_name=?', (company, ))
            c_id = result.fetchall()[0][0]
            result = self.c.execute('SELECT project_name FROM projects WHERE company_id=?', (c_id, ))
            for project in result.fetchall():
                self.project_choices.append(project[0])
            self.project_choices.sort()
            self.refresh_project_option()
            self.project_tk_var.set(' ')
            self.brine_listbox.delete(0, tk.END)
            self.brine_listbox.config(state='disabled')

    def refresh_project_option(self):
        self.project_option['menu'].delete(0, 'end')
        for project in self.project_choices:
            self.project_option['menu'].add_command(label=project, command=partial(self.project_option_setit, project))

    def project_option_setit(self, project):
        self.project_tk_var.set(project)
        self.project_option_command(project)

    def project_option_command(self, *args):
        if args[0] == ' ':
            self.brine_listbox.delete(0, tk.END)
            self.brine_listbox.config(state='disabled')
        else:
            self.brine_listbox.config(state='normal')
            self.brine_listbox_refresh()

    def brine_listbox_refresh(self):
        p_id = self.c.execute('SELECT project_id from projects WHERE project_name=?', (self.project_tk_var.get(), ))
        brines = self.c.execute('SELECT brine_name from brines WHERE project_id=?',
                                (p_id.fetchall()[0][0], )).fetchall()
        brines.sort()
        self.brine_list = [*brines]
        # for i in range(self.brine_list.index('(add new)')):
        for i in range(len(self.brine_list)):
            self.brine_list[i] = brines[i][0]

        self.brine_listbox.delete(0, tk.END)
        for brine in self.brine_list:
            self.brine_listbox.insert(tk.END, brine)

    def browse_brine(self, _):

        brine = self.brine_listbox.get(self.brine_listbox.curselection())

        # if brine == '(add new)':
        #     brine_data = (0, 0, '', '', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, '--', 0, 0., )
        #     self.parent.brine_view.send_button.config(state='normal')
        #     self.parent.brine_view.import_button.config(state='normal')
        # else:
        brine_data_temp = self.c.execute('''SELECT * FROM brines WHERE brine_name=?''', (brine, )).fetchall()[0]
        brine_data = (*brine_data_temp[:-3], '--', *brine_data_temp[-3:], )
        #     self.parent.brine_view.send_button.config(state='disabled')
        #     self.parent.brine_view.import_button.config(state='disabled')

        self.parent.brine_view.name_entry.delete(0, tk.END)
        self.parent.brine_view.name_entry.insert(0, brine_data[2])
        self.parent.brine_view.abbrev_entry.delete(0, tk.END)
        self.parent.brine_view.abbrev_entry.insert(0, brine_data[3])

        ion_names_list = ['lithium', 'sodium', 'potassium', 'magnesium', 'calcium', 'barium', 'strontium', 'ironII',
                          'fluoride', 'chloride', 'bromide', 'bicarbonate', 'carbonate', 'sulfate']

        for j, label in enumerate(self.parent.brine_view.labels):
            ion = label.cget('text')
            entry = self.parent.brine_view.entries[j]
            entry.delete(0, tk.END)
            ind = ion_names_list.index(Brine.ion_names[ion])
            entry.insert(0, str(brine_data[ind + 4]))
            entry.focus_set()
            entry.event_generate('<Return>')

        add_salt_list = ['Na2CO3', 'NaCl']

        for j, label in enumerate(self.parent.brine_view.add_labels):
            add_salt = label.cget('text')
            entry = self.parent.brine_view.add_entries[j]
            entry.delete(0, tk.END)
            ind = add_salt_list.index(add_salt)
            entry.insert(0, str(brine_data[ind + 18]))
            entry.focus_set()
            entry.event_generate('<Return>')


class PrintView(tk.Frame):

    def __init__(self, parent):
        tk.Frame.__init__(self, parent)
        self.parent = parent
        self.print_list_labels = []
        for i in range(18):
            label = tk.Label(self)
            self.print_list_labels.append(label)
            label.pack(side=tk.TOP, padx=2)
        self.print_list_labels[-1].config(width=19)
        self.print_button = tk.Button(self, text='            Print            ',
                                      command=parent.brine_view.print_command)
        self.print_button.config(height=2)
        self.print_button.pack(side=tk.TOP, pady=(27, 5))
        self.config(highlightbackground='black', highlightthickness=1, width=160, height=gui_height-12)
        self.grid(row=0, column=3, rowspan=1, columnspan=1, padx=(0, 10), pady=5)

    def display_print_list(self, print_list: list):

        for label in self.print_list_labels:
            label.config(text=' ')

        for i, item in enumerate(print_list):
            self.print_list_labels[i].config(text=item)


class BrineDatabaseGUI(tk.Tk):

    def __init__(self, export_app=None):
        tk.Tk.__init__(self)
        self.export_app = export_app
        path = db_default_path
        if not isdir('U:\\'):
            path = 'Q' + path[1:]
        self.iconphoto(True, tk.PhotoImage(file=path + 'UEORS logo cropped.png'))
        # self.iconphoto(True, tk.PhotoImage(file=db_default_path + 'UEORS logo cropped.png'))
        self.brine_selector = BrineSelector(self)
        self.brine_view = BrineView(self, NominalBrine())
        self.print_view = PrintView(self)

    def destroy(self):

        if self.brine_selector.conn is not None:
            self.brine_selector.conn.close()

        if self.export_app is not None:
            self.export_app.bd_gui = None

        self.quit()
        tk.Tk.destroy(self)


if __name__ == '__main__':

    app = BrineDatabaseGUI()
    app.title('Brine Database GUI v' + __version__)
    app.geometry(str(gui_width) + 'x' + str(gui_height))
    app.resizable(False, False)

    sys.exit(app.mainloop())
